import openpyxl
import json
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('품목등록 리스트.xlsx', data_only=True)
sheet = wb.active

header = [sheet.cell(2, c).value for c in range(1, 17)]
print("Header columns:", header)

erp_items = []
for r in range(3, sheet.max_row + 1):
    code = sheet.cell(r, 1).value
    name = sheet.cell(r, 2).value
    vendor = sheet.cell(r, 3).value
    spec = sheet.cell(r, 5).value
    unit = sheet.cell(r, 7).value
    use_yn = sheet.cell(r, 13).value
    
    if code and name:
        erp_items.append({
            "code": str(code).strip(),
            "name": str(name).strip(),
            "vendor": str(vendor).strip() if vendor else "",
            "spec": str(spec).strip() if spec else "",
            "unit": str(unit).strip() if unit else "",
            "use_yn": str(use_yn).strip() if use_yn else "YES"
        })

print(f"Total ERP Products parsed: {len(erp_items)}")

# Search for our key product keywords
keywords = ["ANGIO", "안지오", "엔지오", "PENKO", "펜코", "SURGI", "서지", "EZ", "EG", "본드", "이지큐", "DVT", "SLEEVE", "슬리브", "큐어폼", "Cureform", "베리큐어", "Vericure", "스카노스", "GYNE", "가인", "트로카", "Trocar", "Biopsy", "생검", "HYGENT", "하이젠트", "내시경", "소공포"]

print("\n--- MATCHING ERP PRODUCTS FOR KEY SALES ITEMS ---")
matched_catalog = []

for kw in keywords:
    matches = [p for p in erp_items if kw.lower() in p["name"].lower() or kw.lower() in p["spec"].lower() or kw.lower() in p["code"].lower()]
    if matches:
        print(f"\n[Keyword: '{kw}'] -> {len(matches)} items found:")
        for m in matches[:6]:
            print(f"   Code: {m['code']} | Name: {m['name']} | Spec: {m['spec']} | Vendor: {m['vendor']}")

