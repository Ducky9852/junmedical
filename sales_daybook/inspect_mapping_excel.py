import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

subdata_path = r"c:\Users\Ducky98\Desktop\AntiGravity\sales_daybook\subdata\카톡발주_매핑사전_v2.xlsx"
erp_list_path = r"c:\Users\Ducky98\Desktop\AntiGravity\sales_daybook\품목등록 리스트.xlsx"

print("=== 1. Checking 카톡발주_매핑사전_v2.xlsx ===")
wb_sub = openpyxl.load_workbook(subdata_path, data_only=True)
print("Sheets in subdata:", wb_sub.sheetnames)
for sheetname in wb_sub.sheetnames:
    ws = wb_sub[sheetname]
    print(f"\n--- Sheet: {sheetname} (Rows: {ws.max_row}, Cols: {ws.max_column}) ---")
    rows = list(ws.iter_rows(values_only=True))
    if rows:
        print("Header:", rows[0])
        for i, r in enumerate(rows[1:10]):
            print(f"Row {i+1}:", r)

print("\n=== 2. Checking 품목등록 리스트.xlsx ===")
wb_erp = openpyxl.load_workbook(erp_list_path, data_only=True)
print("Sheets in erp_list:", wb_erp.sheetnames)
for sheetname in wb_erp.sheetnames:
    ws = wb_erp[sheetname]
    print(f"\n--- Sheet: {sheetname} (Rows: {ws.max_row}, Cols: {ws.max_column}) ---")
    rows = list(ws.iter_rows(values_only=True))
    if rows:
        print("Header:", rows[0])
        for i, r in enumerate(rows[1:10]):
            print(f"Row {i+1}:", r)
