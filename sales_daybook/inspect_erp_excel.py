import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('품목등록 리스트.xlsx', data_only=True)
print("Sheet Names:", wb.sheetnames)

sheet = wb.active
print(f"Total Rows: {sheet.max_row}, Total Cols: {sheet.max_column}")

header = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
print("\n--- HEADER ---")
for idx, col in enumerate(header):
    if col:
        print(f"Col {idx+1}: {col}")

rows = []
for r in range(2, sheet.max_row + 1):
    row_data = {}
    for c in range(1, len(header) + 1):
        col_name = header[c - 1]
        if col_name:
            row_data[col_name] = sheet.cell(r, c).value
    # check if row is empty
    if any(row_data.values()):
        rows.append(row_data)

print(f"\nTotal Valid Data Rows: {len(rows)}")

print("\n--- SAMPLE 5 ROWS ---")
for i, r in enumerate(rows[:5]):
    print(f"\n[Item {i+1}]")
    for k, v in r.items():
        if v is not None and str(v).strip():
            print(f"  {k}: {v}")
