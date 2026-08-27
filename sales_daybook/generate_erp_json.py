import openpyxl
import json
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('품목등록 리스트.xlsx', data_only=True)
sheet = wb.active

erp_items = []
for r in range(3, sheet.max_row + 1):
    code = sheet.cell(r, 1).value
    name = sheet.cell(r, 2).value
    vendor = sheet.cell(r, 3).value
    spec = sheet.cell(r, 5).value
    unit = sheet.cell(r, 7).value
    use_yn = sheet.cell(r, 13).value
    
    if code and name:
        code_str = str(code).strip()
        name_str = str(name).strip()
        vendor_str = str(vendor).strip() if vendor else ""
        spec_str = str(spec).strip() if spec else ""
        
        # Build keywords
        kws = set()
        kws.add(code_str)
        kws.add(name_str)
        if spec_str:
            kws.add(spec_str)
        for t in re.findall(r'[가-힣]{2,}', name_str + " " + spec_str):
            kws.add(t)
        for t in re.findall(r'[a-zA-Z]{2,}', name_str + " " + spec_str):
            kws.add(t)

        erp_items.append({
            "code": code_str,
            "name": name_str,
            "spec": spec_str,
            "vendor": vendor_str,
            "unit": str(unit).strip() if unit else "EA",
            "category": vendor_str or "의료기기·소모품",
            "keywords": list(kws)
        })

print(f"Total ERP Items Extracted: {len(erp_items)}")

with open('erp_products.json', 'w', encoding='utf-8') as f:
    json.dump(erp_items, f, ensure_ascii=False, indent=2)

print("Saved erp_products.json successfully.")
