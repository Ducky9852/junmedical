import os
import openpyxl
import json
import re
from collections import defaultdict
import sys
import shutil

sys.stdout.reconfigure(encoding='utf-8')

subdata_path = r"c:\Users\Ducky98\Desktop\AntiGravity\sales_daybook\subdata\카톡발주_매핑사전_v2.xlsx"
erp_list_path = r"c:\Users\Ducky98\Desktop\AntiGravity\sales_daybook\품목등록 리스트.xlsx"
erp_list_out_path = r"c:\Users\Ducky98\Desktop\AntiGravity\sales_daybook\품목등록 리스트_updated.xlsx"
erp_json_path = r"c:\Users\Ducky98\Desktop\AntiGravity\sales_daybook\erp_products.json"
sales_db_path = r"c:\Users\Ducky98\Desktop\AntiGravity\sales_daybook\sales_database.json"

print("1. Loading Subdata mapping dictionary...")
wb_sub = openpyxl.load_workbook(subdata_path, data_only=True)
ws_map = wb_sub['매핑표']

code_to_aliases = defaultdict(set)
alias_to_code = {}

for row in ws_map.iter_rows(min_row=2, values_only=True):
    client, alias, code, full_name, valid = row[:5]
    if not code or not alias:
        continue
    code_str = str(code).strip()
    if code_str.endswith('.0'):
        code_str = code_str[:-2]
    alias_str = str(alias).strip()
    
    code_to_aliases[code_str].add(alias_str)
    alias_to_code[alias_str.lower()] = (code_str, str(full_name or '').strip())

print(f"Loaded {len(alias_to_code)} aliases across {len(code_to_aliases)} distinct product codes.")

print("\n2. Updating '품목등록 리스트.xlsx' search keywords (검색창내용)...")
wb_erp = openpyxl.load_workbook(erp_list_path)
ws_erp = wb_erp['품목등록']

updated_excel_rows = 0
for row_idx in range(3, ws_erp.max_row + 1):
    code_cell = ws_erp.cell(row=row_idx, column=1).value
    if not code_cell:
        continue
    code_str = str(code_cell).strip()
    if code_str.endswith('.0'):
        code_str = code_str[:-2]
        
    aliases = code_to_aliases.get(code_str)
    if aliases:
        current_search = ws_erp.cell(row=row_idx, column=11).value
        current_set = set()
        if current_search:
            current_set = set([s.strip() for s in str(current_search).split(',') if s.strip()])
        
        # Merge new aliases
        combined_set = current_set.union(aliases)
        new_search_val = ", ".join(sorted(combined_set))
        ws_erp.cell(row=row_idx, column=11, value=new_search_val)
        updated_excel_rows += 1

try:
    wb_erp.save(erp_list_path)
    print(f"Successfully updated {updated_excel_rows} rows directly into '품목등록 리스트.xlsx'!")
except Exception as e:
    print(f"Could not overwrite directly ({e}), saving to '품목등록 리스트_updated.xlsx'...")
    wb_erp.save(erp_list_out_path)
    try:
        shutil.copyfile(erp_list_out_path, erp_list_path)
        os.remove(erp_list_out_path)
        print("Replaced '품목등록 리스트.xlsx' successfully!")
    except Exception as e2:
        print(f"Keeping '품목등록 리스트_updated.xlsx': {e2}")

print("\n3. Updating erp_products.json with aliases and building fast search master...")
with open(erp_json_path, 'r', encoding='utf-8') as f:
    erp_products = json.load(f)

for p in erp_products:
    p_code = str(p['code']).strip()
    if p_code in code_to_aliases:
        kw_set = set(p.get('keywords', []))
        for al in code_to_aliases[p_code]:
            kw_set.add(al.lower())
            for word in al.split():
                if len(word) >= 2:
                    kw_set.add(word.lower())
        p['keywords'] = list(kw_set)
        p['aliases'] = list(code_to_aliases[p_code])

with open(erp_json_path, 'w', encoding='utf-8') as f:
    json.dump(erp_products, f, ensure_ascii=False, indent=2)

print(f"Successfully updated erp_products.json ({len(erp_products)} products).")

print("\n4. Re-evaluating sales_database.json logs with enhanced mapping rules...")
with open(sales_db_path, 'r', encoding='utf-8') as f:
    sales_db = json.load(f)

# Build search lookup table
prod_by_code = {p['code']: p for p in erp_products}

remap_count = 0
for log in sales_db.get('activity_logs', []):
    title = log.get('title', '')
    note = log.get('note', '')
    combined_text = f"{title} {note}".lower()
    
    # Check if current product is generic or missing
    current_prods = log.get('products', [])
    current_prod_name = current_prods[0] if current_prods else ''
    
    matched_item = None
    
    # 1. Match from alias dictionary
    for al, (code, full_name) in alias_to_code.items():
        if len(al) >= 2 and al in combined_text:
            if code in prod_by_code:
                matched_item = prod_by_code[code]
                break

    # 2. Match from specific rules
    if not matched_item:
        rules = [
            (r'dvt|슬리브|암슬리브|스타킹|허벅지', 'MBH02'),
            (r'penko|펜코|서지패드|서지소드|보비', 'PK-CGP202S-TB'),
            (r'베리큐어|vericure', 'MD-BR-001'),
            (r'gyne|가인콜라|gnc', 'GNC2505D'),
            (r'소공포|med2078', 'MED2078'),
            (r'angio|엔지오|안지오|angiocath', '382412'),
            (r'trocar|트로카', '101.011A'),
            (r'biopsy|생검|펀치바이옵시|바이옵시', '045-7301'),
            (r'내시경|endo', 'GEPL-F1'),
            (r'좌욕기|좌욕기.*필터|필터.*교체', 'FIL-ZY01'),
            (r'bt350|bt-350|태아감시장치', 'BT-350-DEV'),
            (r'튤립|듀얼튤립', 'TL-DUAL-01'),
            (r'루미|rumi', 'UMB678'),
            (r'스카노스|scarnos', 'SCAR-GEL-01'),
            (r'탑매트', 'ch1023'),
            (r'이지포트|easyport', '0-21-0040'),
            (r'젠타큐|genta', '070-1301')
        ]
        for pat, code in rules:
            if re.search(pat, combined_text, re.I):
                if code in prod_by_code:
                    matched_item = prod_by_code[code]
                    break

    if matched_item:
        new_prod_name = matched_item['name']
        if current_prod_name != new_prod_name and ('일반' in current_prod_name or '미지정' in current_prod_name or not current_prod_name or 'PROD' in current_prod_name):
            log['products'] = [new_prod_name]
            log['product_code'] = matched_item['code']
            remap_count += 1
        elif not log.get('product_code'):
            log['product_code'] = matched_item['code']

print(f"Refined {remap_count} activity logs with accurate ERP mapping codes.")

with open(sales_db_path, 'w', encoding='utf-8') as f:
    json.dump(sales_db, f, ensure_ascii=False, indent=2)

js_path = os.path.join(os.path.dirname(sales_db_path), "sales_database.js")
with open(js_path, 'w', encoding='utf-8') as f:
    f.write("window.SALES_DB = " + json.dumps(sales_db, ensure_ascii=False, indent=2) + ";\n")

print("Saved updated sales_database.json and sales_database.js!")
