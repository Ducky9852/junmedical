import openpyxl
import json
import sys
import re
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

# 1. Load ERP Item Master with EDI (보험코드)
wb = openpyxl.load_workbook('품목등록 리스트.xlsx', data_only=True)
sheet = wb.active

erp_catalog = []
erp_by_code = {}

for r in range(3, sheet.max_row + 1):
    code = sheet.cell(r, 1).value
    name = sheet.cell(r, 2).value
    vendor = sheet.cell(r, 3).value
    edi = sheet.cell(r, 4).value
    spec = sheet.cell(r, 5).value
    unit = sheet.cell(r, 7).value
    use_yn = sheet.cell(r, 13).value
    
    if code and name:
        code_str = str(code).strip()
        name_str = str(name).strip()
        vendor_str = str(vendor).strip() if vendor else ""
        edi_str = str(edi).strip() if edi else ""
        spec_str = str(spec).strip() if spec else ""
        
        full_display_name = f"{name_str} ({spec_str})" if spec_str else name_str
        
        # Build keywords
        kws = set()
        kws.add(code_str)
        kws.add(name_str)
        if spec_str:
            kws.add(spec_str)
        if edi_str:
            kws.add(edi_str)
        for t in re.findall(r'[가-힣]{2,}', name_str + " " + spec_str):
            kws.add(t)
        for t in re.findall(r'[a-zA-Z]{2,}', name_str + " " + spec_str):
            kws.add(t)

        item = {
            "id": code_str,
            "code": code_str,
            "name": full_display_name,
            "raw_name": name_str,
            "spec": spec_str,
            "edi": edi_str,
            "vendor": vendor_str,
            "unit": str(unit).strip() if unit else "EA",
            "category": vendor_str or "의료기기·소모품",
            "keywords": list(kws)
        }
        erp_catalog.append(item)
        erp_by_code[code_str] = item

print(f"Total ERP Catalog Loaded: {len(erp_catalog)} items")

# 2. Precision Mapping Rules (Exact ERP code mapping based on product names in Notion)
EXACT_MAPPING_RULES = [
    # Angio Kit Series
    (r'PR03|PR03\s*\(Adv\.\)|Adv03|ANGIO.*PR03', 'ST-ANG-PR03'),
    (r'PR04|시술용04|Adv04|ANGIO.*PR04', 'ST-ANG-PR04'),
    (r'PR01|Basic01|ANGIO.*PR01', 'ST-ANG-PR01'),
    (r'PR07|ANGIO.*PR07', 'ST-ANG-PR07'),
    (r'PR10|ANGIO.*PR10', 'ST-ANG-PR10'),
    (r'ANGIO|안지오|엔지오|시술용\s*키트', 'ST-ANG-PR03'),  # Default to PR03 Kit instead of needle!

    # PENKO Surgi Pad & Sword
    (r'PK-CGP202S|보비|플레이트|펜코\s*BV', 'PK-CGP202S-TB'),
    (r'PK-11DMS|STRIP\s*SURGI\s*SWORD', 'PK-11DMS'),
    (r'PK-15DM|DF\s*SURGI\s*SWORD|서지소드', 'PK-15DM02'),
    (r'서포트커버|PKSCB', 'PKSCB-M'),
    (r'PENKO|펜코|서지패드|서포트\s*플레이트', 'PK-CGP202S-TB'),

    # EZ-Q Bond (조직접착제)
    (r'EG-Q\s*Bond\s*N\s*Mini|본드\s*미니', 'WB-EGQ-S-02'),
    (r'EG-Q|EZ-Q|EzQbond|이지큐\s*본드|본드', 'WB-EGQ-A--02'),

    # Scarnos Gel (스카노스겔)
    (r'스카노스|Scarnos', '194573'),

    # DVT SLEEVE
    (r'발형|MBH04-02', 'MBH04-02'),
    (r'종아리|MBH04-01', 'MBH04-01'),
    (r'허벅지|MBH02|DVT|슬리브|스타킹|암슬리브', 'MBH02'),

    # Vericure (베리큐어)
    (r'스프레이|MD-BR-003', 'MD-BR-003'),
    (r'베리큐어|Vericure|로션|MD-BR-001', 'MD-BR-001'),

    # GYNE COLLA (가인콜라)
    (r'GYNE|가인콜라|GNC', 'GNC2505D'),

    # 소공포
    (r'70\*70|MED5741', 'MED5741'),
    (r'소공포|50\*50|MED2078', 'MED2078'),

    # Trocar (트로카)
    (r'Crown.*Blade|121-01825', '121-01825'),
    (r'트로카|Trocar|101\.011A', '101.011A'),

    # Biopsy (생검기구)
    (r'EN-Shot|Biopsy|생검', '045-7301')
]

# 3. Update Sales Database
with open('sales_database.json', 'r', encoding='utf-8') as f:
    sales_db = json.load(f)

updated_pipeline = []
deals_by_key = {}

for deal in sales_db['pipeline']:
    prod_name = deal['product_name']
    matched_erp = None
    
    # Try exact rule mapping
    for pat, erp_code in EXACT_MAPPING_RULES:
        if re.search(pat, prod_name, re.I) or re.search(pat, deal.get('latest_note', ''), re.I):
            if erp_code in erp_by_code:
                matched_erp = erp_by_code[erp_code]
                break
                
    if not matched_erp:
        for erp in erp_catalog:
            if erp['raw_name'] in prod_name or prod_name in erp['name']:
                matched_erp = erp
                break

    if matched_erp:
        deal['product_id'] = matched_erp['id']
        deal['product_name'] = matched_erp['name']
        deal['product_category'] = matched_erp['category']
        deal['vendor'] = matched_erp['vendor']
        deal['erp_code'] = matched_erp['id']
        deal['edi'] = matched_erp.get('edi', '')
    else:
        deal['erp_code'] = deal['product_id']
        deal['edi'] = ''
        
    deal_key = f"{deal['hospital']}___{deal['product_id']}"
    if deal_key not in deals_by_key:
        deals_by_key[deal_key] = deal
        updated_pipeline.append(deal)
    else:
        existing = deals_by_key[deal_key]
        if deal['last_date'] > existing['last_date']:
            existing.update(deal)

print(f"Updated Pipeline Deals: {len(updated_pipeline)}")

# Active ERP Products to appear top in pills
active_prod_ids = set([d['product_id'] for d in updated_pipeline])
top_erp_products = [p for p in erp_catalog if p['id'] in active_prod_ids]
other_erp_products = [p for p in erp_catalog if p['id'] not in active_prod_ids]
final_product_catalog = top_erp_products + other_erp_products

# Update Stats
won_count = len([d for d in updated_pipeline if d['status'] == '도입완료·납품'])
demo_count = len([d for d in updated_pipeline if d.get('demo_info') and d['demo_info']['status'] == '평가진행중'])
as_count = len([d for d in updated_pipeline if d.get('as_info') and d['as_info']['status'] == '접수/진행중'])
lost_count = len([d for d in updated_pipeline if d['status'] == '영업실패·보류'])
eval_count = len([d for d in updated_pipeline if d['status'] == '데모·샘플평가'])
progress_count = len([d for d in updated_pipeline if d['status'] in ['제품소개·영업중', '견적·의사결정', '관계관리·접촉']])

sales_db['products'] = final_product_catalog
sales_db['pipeline'] = updated_pipeline
sales_db['stats'] = {
    "total_logs": len(sales_db['activity_logs']),
    "total_hospitals": len(sales_db['hospitals']),
    "total_deals": len(updated_pipeline),
    "total_erp_items": len(final_product_catalog),
    "active_demos": demo_count,
    "active_as": as_count,
    "won_deals": won_count,
    "lost_deals": lost_count,
    "eval_deals": eval_count,
    "progress_deals": progress_count
}

# Save updated JSON & JS database
with open('sales_database.json', 'w', encoding='utf-8') as f:
    json.dump(sales_db, f, ensure_ascii=False, indent=2)

with open('sales_database.js', 'w', encoding='utf-8') as f:
    f.write("window.SALES_DB = " + json.dumps(sales_db, ensure_ascii=False, indent=2) + ";\n")

with open('erp_products.json', 'w', encoding='utf-8') as f:
    json.dump(erp_catalog, f, ensure_ascii=False, indent=2)

print("\n=== Precision ERP Mapping Complete ===")
# Verify Yuseong Sun Hospital item
yuseong_deal = next((d for d in updated_pipeline if '유성선' in d['hospital']), None)
if yuseong_deal:
    print(f"Verified Yuseong Sun Hospital: {yuseong_deal['hospital']} -> Code: {yuseong_deal['product_id']}, Name: {yuseong_deal['product_name']}, EDI: {yuseong_deal.get('edi')}")
