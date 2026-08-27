import os
import json
import openpyxl
import urllib.request
import ssl
import sys

sys.stdout.reconfigure(encoding='utf-8')

SUPABASE_URL = "https://hkvguhttmxclyaeskznk.supabase.co"
SUPABASE_KEY = "sb_publishable_qZvInHl5ds9HXTJ_cMF7-g_0P-SefMJ"

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates"
}

cur_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(cur_dir, "품목등록 리스트_updated.xlsx")
erp_json_path = os.path.join(cur_dir, "erp_products.json")
db_json_path = os.path.join(cur_dir, "sales_database.json")
db_js_path = os.path.join(cur_dir, "sales_database.js")

print("1. Reading Excel:", excel_path)
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

# Row 2 is header:
# Col 1: 품목코드, Col 2: 품목명, Col 3: 구매처명, Col 4: 보험코드, Col 5: 규격명
# Col 6: 입고단가, Col 7: 단위, Col 8: EA/BOX, Col 9: 출고단가, Col 11: 검색창내용
all_products = []

for r in range(3, ws.max_row + 1):
    code = ws.cell(row=r, column=1).value
    name = ws.cell(row=r, column=2).value
    if not code or not str(code).strip():
        continue
        
    code = str(code).strip()
    name = str(name).strip() if name else code
    vendor = str(ws.cell(row=r, column=3).value or '').strip()
    edi = str(ws.cell(row=r, column=4).value or '').strip()
    spec = str(ws.cell(row=r, column=5).value or '').strip()
    price_in = ws.cell(row=r, column=6).value or 0
    unit = str(ws.cell(row=r, column=7).value or 'EA').strip()
    price_out = ws.cell(row=r, column=9).value or 0
    search_terms = str(ws.cell(row=r, column=11).value or '').strip()
    
    try:
        price_in = float(price_in) if price_in else 0
    except:
        price_in = 0
    try:
        price_out = float(price_out) if price_out else 0
    except:
        price_out = 0

    keywords = [k for k in [code, name, vendor, edi, spec, search_terms] if k]
    aliases = [k.strip() for k in search_terms.split(',') if k.strip()] if search_terms else []
    
    all_products.append({
        "id": code,
        "code": code,
        "name": name,
        "spec": spec,
        "vendor": vendor,
        "category": "일반의료기기",
        "edi": edi,
        "unit": unit,
        "price_in": price_in,
        "price_out": price_out,
        "keywords": keywords,
        "aliases": aliases
    })

print(f"Extracted {len(all_products)} products from Excel!")

# Check ST-ANG-PR03
pr03 = next((p for p in all_products if "PR03" in p['code']), None)
print(f"Found PR03 in Excel parsed: {pr03}")

# Add Parent Groups at top
from setup_penko_sword import PARENT_GROUP_PENKO_SWORD, PENKO_SWORD_SKUS
from setup_sejong_sheet import PARENT_GROUP_PRODUCT, SEJONG_SHEET_SKUS

existing_codes = {p['code'] for p in all_products}

# Prepend groups
groups_to_add = [PARENT_GROUP_PENKO_SWORD, PARENT_GROUP_PRODUCT]
for g in reversed(groups_to_add):
    if g['code'] not in existing_codes:
        all_products.insert(0, g)

# Also ensure all SKUs are in
for sku_list in [PENKO_SWORD_SKUS, SEJONG_SHEET_SKUS]:
    for sku in sku_list:
        if sku['code'] not in existing_codes:
            sku_obj = dict(sku)
            sku_obj['id'] = sku['code']
            all_products.insert(2, sku_obj)
            existing_codes.add(sku['code'])

print(f"Total products with groups and SKUs: {len(all_products)}")

# 2. Write to erp_products.json
with open(erp_json_path, 'w', encoding='utf-8') as f:
    json.dump(all_products, f, ensure_ascii=False, indent=2)
print("Saved erp_products.json successfully!")

# 3. Update sales_database.json & js
with open(db_json_path, 'r', encoding='utf-8') as f:
    db = json.load(f)

db['products'] = all_products
with open(db_json_path, 'w', encoding='utf-8') as f:
    json.dump(db, f, ensure_ascii=False, indent=2)

with open(db_js_path, 'w', encoding='utf-8') as f:
    # Also attach ERP_PRODUCTS_MASTER directly in window for instant fast access
    f.write("window.ERP_PRODUCTS_MASTER = " + json.dumps(all_products, ensure_ascii=False) + ";\n")
    f.write("window.SALES_DB = " + json.dumps(db, ensure_ascii=False) + ";\n")
print("Saved sales_database.json & sales_database.js with full 4,050+ master products!")

# 4. Sync full products to Supabase
print("\n--- Syncing Full 4,050+ Products to Supabase erp_products ---")
supabase_erp = []
for p in all_products:
    supabase_erp.append({
        "code": p["code"],
        "name": p["name"],
        "spec": p.get("spec", ""),
        "vendor": p.get("vendor", ""),
        "category": p.get("category", "일반의료기기"),
        "keywords": p.get("keywords", []),
        "aliases": p.get("aliases", []),
        "price_in": p.get("price_in", 0),
        "price_out": p.get("price_out", 0)
    })

for i in range(0, len(supabase_erp), 150):
    batch = supabase_erp[i:i+150]
    req = urllib.request.Request(f"{SUPABASE_URL}/rest/v1/erp_products", data=json.dumps(batch).encode('utf-8'), headers=HEADERS, method='POST')
    try:
        with urllib.request.urlopen(req, context=ctx) as res:
            pass
    except Exception as e:
        print(f"Supabase batch {i} error:", e)

print(f"Synced {len(supabase_erp)} products to Supabase cloud!")
